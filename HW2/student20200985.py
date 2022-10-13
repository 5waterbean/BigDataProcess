#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

scores = []
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		scores.append(sum_v)
	row_id += 1

scores.sort(reverse=True)

count = row_id - 2
row_id = 1
cntA = 0
cntB = 0
cntC = 0
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		if score > scores[int(count*0.3)]:
			cntA += 1
		elif score > scores[int(count*0.7)]:
			cntB += 1
		else:
			cntC += 1
	row_id += 1

grade = ""
row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		if score > scores[int(cntA/2)]:
			grade = "A+"
		elif score > scores[cntA]:
			grade = "A"
		elif score > scores[cntA+int(cntB/2)]:
			grade = "B+"
		elif score > scores[cntA+cntB]:
			grade = "B"
		elif score > scores[cntA+cntB+int(cntC/2)]:
			grade = "C+"
		else:
			grade = "C"
		ws.cell(row = row_id, column = 8).value = grade
	row_id += 1

wb.save("student.xlsx")
